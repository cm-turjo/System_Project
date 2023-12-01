

import openpyxl

workbook = openpyxl.load_workbook("D:/Academic/System Final/Solution/11111names_and_emails.xlsx")

sheet = workbook.active

total_rows = sheet.max_row

# Iterate over rows starting from the 2nd row
for row_index in range(2, total_rows + 1):
    # Get the value from the first column (column index 1)
    cell_value = sheet.cell(row=row_index, column=1).value
    print(cell_value)
    
    cell_value = sheet.cell(row=row_index, column=4).value
    print(cell_value)

# Close the workbook
workbook.close()