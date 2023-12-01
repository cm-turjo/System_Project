from openpyxl import load_workbook
from openpyxl.formula import Tokenizer, Token

class Teacher:
    def __init__(self, name, designation, dept, question_paper_setter, examiners_class_tests, examiners_sessional_classes,
                 script_scrutinizer, tabulation, typing_and_drawing, central_viva_voce,
                 student_advising, seminar, thesis_progress_defense, final_grade_sheet_verification,
                 list_of_duty, sessional_course, theory_course):
        self.name = name
        self.designation = designation
        self.dept = dept
        self.question_paper_setter = int(question_paper_setter)
        self.examiners_class_tests = int(examiners_class_tests)
        self.examiners_sessional_classes = int(examiners_sessional_classes)
        self.script_scrutinizer = int(script_scrutinizer)
        self.tabulation = int(tabulation)
        self.typing_and_drawing = int(typing_and_drawing)
        self.central_viva_voce = int(central_viva_voce)
        self.student_advising = int(student_advising)
        self.seminar = int(seminar)
        self.thesis_progress_defense = int(thesis_progress_defense)
        self.final_grade_sheet_verification = int(final_grade_sheet_verification)
        self.list_of_duty = int(list_of_duty)
        self.sessional_course = sessional_course
        self.theory_course = theory_course

    def to_dict(self):
        return {
            'Name': self.name,
            'Designation': self.designation,
            'Dept': self.dept,
            'Question Paper Setter': self.question_paper_setter,
            'Examiners of Class Tests': self.examiners_class_tests,
            'Examiners of Sessional Classes': self.examiners_sessional_classes,
            'Script Scrutinizer': self.script_scrutinizer,
            'Tabulation': self.tabulation,
            'Typing and Drawing': self.typing_and_drawing,
            'Central Viva-Voce': self.central_viva_voce,
            'Student Advising': self.student_advising,
            'Seminar': self.seminar,
            'Thesis Progress Defense': self.thesis_progress_defense,
            'Final Grade Sheet Verification': self.final_grade_sheet_verification,
            'List of Duty': self.list_of_duty,
            'Sessional Course': self.sessional_course,
            'Theory Course': self.theory_course,
        }



teachers_array = []

from docx import Document
from docx.shared import Pt
from docx.enum.text import WD_UNDERLINE
from googletrans import Translator
import apicall1 as apc



TName = []


designation_map = {}
dept_map = {}
question_paper_setter_map = {}
examiners_class_tests_map = {}
examiners_sessional_classes_map = {}
script_scrutinizer_map = {}
tabulation_map = {}
typing_and_drawing_map = {}
central_viva_voce_map = {}
student_advising_map = {}
seminar_map = {}
thesis_progress_defense_map = {}
final_grade_sheet_verification_map = {}
list_of_duty_map = {}
sessional_course_map={}
sessional_course_Cradit_map={}
theory_course_map={}



px = {}



bangla_teachers_name = {
    'Dr. Md. Shahjahan' : 'ড. মো. শাহজাহান',
    'Dr. Sk. Md. Masudul Ahsan' : 'ড. এস কে মো. মাসুদুল আহসান',
    'Dr. M. M. A. Hashem' : 'ড. এম এম এ হাশেম',
    'Dr. K. M. Azharul Hasan' : 'ড. কে এম আজহারুল হাসান',
    'Dr. Kazi Md. Rokibul Alam' : 'ড. কাজী মো. রকিবুল আলম',
    'Dr. Muhammad Sheikh Sadi' : 'ড. মোহাম্মদ শেখ সাদী',
    'Mr. Md. Abdus Salim Mollah' : 'মো. আব্দুস সেলিম মোল্লা',
    'Dr. Sk. Imran Hossain' : 'ড. এস কে ইমরান হোসেন',
    'Mr. Abdul Aziz' : 'আব্দুল আজিজ',
    'Mr. Sunanda Das' : 'সুনন্দ দাস',
    'Mrs. Nazia Jahan Khan Chowdhury' : 'নাজিয়া জাহান খান চৌধুরী',
    'Mrs. Dola Das' : 'দোলা দাস',
    'Mr. S. M. Taslim Uddin Raju' : 'এস এম তসলিম উদ্দিন রাজু',
    'Mr. Md. Motaleb Hossen Manik' : 'মো. মোত্তালেব হোসেন মানিক',
    'Mr. Md. Ahsan Habib' : 'মো. আহসান হাবীব',
    'Mr. Md. Shahidul Salim' : 'মো. শাহীদুল সেলিম',
    'Ms. Dipannita Biswas' : 'দীপান্বিতা বিশ্বাস',
    'Mr. Md. Repon Islam' : 'মো. রিপন ইসলাম',
    'Mr. Md. Sakhawat Hossain' : 'মো. শাখওয়াত হোসেন',
    'Mr. Md. Nazirulhasan Shawon' : 'মো. নাজিরুল হাসান শাওন',
    'Md. Badiuzzaman Shuvo' : 'মো. বদীউজ্জামান শুভ',
    'Most. Kaniz Fatema Isha' : 'মোছা: কানিজ ফাতেমা ইশা',
    'Mr. Farhan Sadaf' : 'ফারহান সাদাফ',
    'Mr. Safin Ahmmed' : 'শাফিন আহমেদ',
    'Mr. Argha Chandra Dhar' : 'অর্ঘ্য চন্দ্র ধর',
    'Mr. Md Mehrab Hossain Opi' : 'মো. মেহরাব হোসেন অপি',
    'Dr. Subrata Talapatra' : 'ড. সুব্রত তালপত্র',
    'Mr. Ridwan Mustofa' : 'রিদওয়ান মোস্তফা',
    'Mr. Jahid Hasan Ashik' : 'জাহিদ হাসান আশিক',
    'Dr. A. R. M. Jalal Uddin Jamali' : 'ড. এ আর এম জালাল উদ্দিন জামালী',
    'Dr. Md. Alhaz Uddin' : 'ড. মো. আলহাজ উদ্দিন',
    'Dr. S. M. Rabiul Alam' : 'ড. এস এম রবিউল আলম',
    'Dr. Md. Hasanuzzaman' : 'ড. মো. হাসানুজ্জামান',
    'Mr. Md. Hasibul Haque' : 'মো. হাসিবুল হক'
    
}


bangla_teachers_designation ={
    'Dean' : 'ডীন',
    'Professor' : 'অধ্যাপক',
    'Associate Professor' : 'সহযোগী অধ্যাপক',
    'Assistant Professor' : 'সহকারী অধ্যাপক',
    'Lecturer' : 'প্রভাষক'
    
}


def is_underlined(run):
    """
    Check if a run contains underlined text.
    """
    return run.font.underline != WD_UNDERLINE.NONE

def find_table_heading(table):
    """
    Find the heading of a table by looking at the preceding paragraphs.
    """
    for paragraph in table._element.getprevious():
        if paragraph.tag.endswith("p"):
            if any(is_underlined(run) for run in paragraph.runs):
                return paragraph.text.strip()
    return None

def translate_to_bengali(english_name):
    translator = Translator()
    translation = translator.translate(english_name, src='en', dest='bn')
    return translation.text


def size(mapp):
    ck = 0
    for teacher in TName:
        # Assuming TName is a list of teachers, you might want to iterate over them
        cccc = mapp.get(teacher, 0)
        if cccc != 0:
        #   print(teacher)
        #   print(mapp[teacher])
          ck += 1
        

    return ck

def print_tables(doc):
    for i, table in enumerate(doc.tables, start=1):
        # Extracting heading from the preceding paragraphs
        #heading = find_table_heading(table)
        # Printing table heading
        #print(f"\nTable {i} Heading: {heading}")
        # Printing rows and values
        #print("-------------------Table :---------------------------"+str(i))

        if i == 1:
            #print("-------------------Table : 1---------------------------")
            r = 1
            for row in table.rows:
                data = "demo"
                data = table.rows[r].cells[1].text.strip()
                # print(data)

                data2 = "demo2"
                data2 = table.rows[r].cells[2].text.strip()
                # print(data2)

                if data2 == "Total =":
                    continue
                if data == "Total =":
                    continue
                result = data2.split(',')

                d2=result[0]
                d3=result[1]


                TName.append(data)
                designation_map[data] = d2
                dept_map[data] = d3

                r = r + 1;


        if i == 2:

            #print("-------------------Table : 2---------------------------")
            heading = find_table_heading(table)
            r = 1
            dm=0
            for row in table.rows:
                if dm==0:
                  dm=1
                  continue
                data = "demo"
                data = table.rows[r].cells[1].text.strip()
                #print(data)

                data2 = "demo2"
                data2 = table.rows[r].cells[3].text.strip()
                #print(data2)

                if data2 == "Total =":
                    continue
                if data == "Total =":
                    continue
                #data = translate_to_bengali[data]
                question_paper_setter_map[data] = data2
                #print("Mumdu:"+str(r))

                #print(question_paper_setter_map[data])
                #print("------------------Mumdu:----------------")


                r = r + 1;
            #print("-------------------ck-----------------")

        if i == 3:

            #print("-------------------Table : 3---------------------------")
            heading = find_table_heading(table)
            r = 1

            dm=0
            for row in table.rows:
                if dm==0:
                  dm=1
                  continue
                data = "demo"
                data = table.rows[r].cells[1].text.strip()
                # print(data)

                data2 = "demo2"
                data2 = table.rows[r].cells[2].text.strip()

                data3 = "demo3"
                data3 = table.rows[r].cells[0].text.strip()

                # print(data2)

                if data2 == "Total =":
                    continue
                if data == "Total =":
                    continue
                if data3 == "Total =":
                    continue

                examiners_class_tests_map[data] = data2
                theory_course_map[data] = data3

                r = r + 1;

                #print(f"{data} is {data2}")

        if i == 4:
           # print("-------------------Table : 4---------------------------")
            heading = find_table_heading(table)
            r = 1

            dm=0
            for row in table.rows:
                if dm==0:
                  dm=1
                  continue


                data = "demo"
                data = table.rows[r].cells[1].text.strip()
                #print(data)

                data2 = "demo2"
                data2 = table.rows[r].cells[2].text.strip()
                data3 = 0.00
                data3 = float(table.rows[r].cells[3].text.strip())
                #print(data3)


                # print(data2)
                data4 = "demo3"
                data4 = table.rows[r].cells[0].text.strip()
                

                if data2 == "Total =":
                    continue
                if data == "Total =":
                    continue
                examiners_sessional_classes_map[data] = data2
                sessional_course_map[data]=data4
                sessional_course_Cradit_map[data4]=data3
               # print("Data 1 is ")
                #print(data)
                
                # #print("Data 2 is ")
                # # print(data2)
                # # print("Data 4 is ")
                # # print(data4)
                #print("Sessional Course Map ")
                #print(sessional_course_map[data])
                # print("Examiners Sessional Classes is ")
                # print(sessional_course_Cradit_map[data4])

                r = r + 1;
                #print("-------------------Table : 4---------ENDDDD------------------")


        if i == 5:
            #print("-------------------Table : 5---------------------------")
            heading = find_table_heading(table)
            r = 1


            dm=0
            for row in table.rows:
                if dm==0:
                  dm=1
                  continue



                data = "demo"
                data = table.rows[r].cells[0].text.strip()
                # print(data)

                data2 = "demo2"
                data2 = table.rows[r].cells[1].text.strip()
                # print(data2)

                if data2 == "Total =":
                    continue
                if data == "Total =":
                    continue

                result = data2.split('=')

                d2=int(result[1])


                script_scrutinizer_map[data] = d2

                r = r + 1;


        if i == 6:
            #print("-------------------Table : 6---------------------------")
            heading = find_table_heading(table)
            r = 1


            dm=0
            for row in table.rows:
                if dm==0:
                  dm=1
                  continue


                data = "demo"
                data = table.rows[r].cells[1].text.strip()
                # print(data)

                data2 = "demo2"
                data2 = table.rows[r].cells[2].text.strip()
                # print(data2)

                if data2 == "Total =":
                    continue
                if data == "Total =":
                    continue
                tabulation_map[data] = data2

                r = r + 1;


        if i == 7:
            heading = find_table_heading(table)
            r = 1


            dm=0
            for row in table.rows:
                if dm==0:
                  dm=1
                  continue



                data = "demo"
                data = table.rows[r].cells[0].text.strip()
                # print(data)

                data2 = "demo2"
                data2 = table.rows[r].cells[1].text.strip()
                # print(data2)

                if data2 == "Total =":
                    continue
                if data == "Total =":
                    continue
                typing_and_drawing_map[data] = data2

                r = r + 1;




        if i == 8:
            heading = find_table_heading(table)
            r = 1



            dm=0
            for row in table.rows:
                if dm==0:
                  dm=1
                  continue



                data = "demo"
                data = table.rows[r].cells[1].text.strip()
                # print(data)

                data2 = "demo2"
                data2 = table.rows[r].cells[2].text.strip()
                # print(data2)

                if data2 == "Total =":
                    continue
                if data == "Total =":
                    continue
                central_viva_voce_map[data] = data2

                r = r + 1;




        if i == 9:
            heading = find_table_heading(table)
            r = 1



            dm=0
            for row in table.rows:
                if dm==0:
                  dm=1
                  continue



                data = "demo"
                data = table.rows[r].cells[1].text.strip()
                # print(data)

                data2 = "demo2"
                data2 = table.rows[r].cells[2].text.strip()
                # print(data2)

                if data2 == "Total =":
                    continue
                if data == "Total =":
                    continue
                student_advising_map[data] = data2

                r = r + 1;


        if i == 10:
            heading = find_table_heading(table)
            r = 1



            dm=0
            for row in table.rows:
                if dm==0:
                  dm=1
                  continue

                data = "demo"
                data = table.rows[r].cells[1].text.strip()
                # print(data)

                data2 = "demo2"
                data2 = table.rows[r].cells[2].text.strip()
                # print(data2)

                if data2 == "Total =":
                    continue
                if data == "Total =":
                    continue
                seminar_map[data] = data2

                r = r + 1;


        if i == 11:
            heading = find_table_heading(table)
            r = 1


            dm=0
            for row in table.rows:
                if dm==0:
                  dm=1
                  continue



                data = "demo"
                data = table.rows[r].cells[1].text.strip()
                # print(data)

                data2 = "demo2"
                data2 = table.rows[r].cells[2].text.strip()
                # print(data2)

                if data2 == "Total =":
                    continue
                if data == "Total =":
                    continue
                thesis_progress_defense_map[data] = data2

                r = r + 1;


        if i == 12:
            heading = find_table_heading(table)
            r = 1


            dm=0
            for row in table.rows:
                if dm==0:
                  dm=1
                  continue



                data = "demo"
                data = table.rows[r].cells[1].text.strip()
                # print(data)

                data2 = "demo2"
                data2 = table.rows[r].cells[2].text.strip()
                # print(data2)

                if data2 == "Total =":
                    continue
                if data == "Total =":
                    continue

                #result = data2.split('(')
                #d2=int(result[0])


                final_grade_sheet_verification_map[data] = data2
                # print("Final Grade Sheet Verification: ")
                # print(data)
                # print(data2)
                # print(final_grade_sheet_verification_map[data])
                # print("...............")

                r = r + 1;

        if i == 13:
            heading = find_table_heading(table)
            r = 1


            dm=0
            for row in table.rows:
                if dm==0:
                  dm=1
                  continue



                data = "demo"
                data = table.rows[r].cells[1].text.strip()
                # print(data)

                data2 = "demo2"
                data2 = table.rows[r].cells[2].text.strip()
                # print(data2)

                if data2 == "Total =":
                    continue
                if data == "Total =":
                    continue
                list_of_duty_map[data] = data2

                r = r + 1;


def create_teacher(i):
    name = TName[i]
    designation = designation_map[name]
    dept = dept_map[name]


    question_paper_setter = question_paper_setter_map.get(name, 0)
    examiners_class_tests = examiners_class_tests_map.get(name, 0)
    examiners_sessional_classes = examiners_sessional_classes_map.get(name, 0)
    script_scrutinizer = script_scrutinizer_map.get(name, 0)
    tabulation = tabulation_map.get(name, 0)
    typing_and_drawing = typing_and_drawing_map.get(name, 0)
    central_viva_voce = central_viva_voce_map.get(name, 0)
    student_advising = student_advising_map.get(name, 0)
    seminar = seminar_map.get(name, 0)
    thesis_progress_defense = thesis_progress_defense_map.get(name, 0)
    final_grade_sheet_verification = final_grade_sheet_verification_map.get(name, 0)
    list_of_duty = list_of_duty_map.get(name, 0)
    sessional_course = sessional_course_map.get(name, "Null")
    theory_course = theory_course_map.get(name, "Null")
    
    # print("Showing The values.........................................")
    
    # print(name)
                
    #             #print("Data 2 is ")
    #             # print(data2)
    #             # print("Data 4 is ")
    #             # print(data4)
    # print("Sessional Course Map ")
    # print(sessional_course_map["Md. Badiuzzaman Shuvo"])
    # print(".............................................")


    # Creating and returning the Teacher instance
    return Teacher(name, designation, dept, question_paper_setter, examiners_class_tests,
                   examiners_sessional_classes, script_scrutinizer, tabulation, typing_and_drawing,
                   central_viva_voce, student_advising, seminar, thesis_progress_defense,
                   final_grade_sheet_verification, list_of_duty, sessional_course, theory_course)



def excel_add(cteacher):
    workbook = load_workbook('D:\\Academic\\System Final\\imransirdemo.xlsx')
    name_string = "নাম: "
    podobi_string = "পদবী: "
    bivag_string = "বিভাগ/শাখা: "
    # Select the active sheet
    sheet = workbook.active
    
    # Writing Name to the excel Sheet Demo( Name: Dr Sheikh Imran Hossain)
    merged_cell = sheet[ 'A3:C3' ]
    new_value = cteacher.name
    new_value = bangla_teachers_name[cteacher.name]
    #print("New Value is : ", new_value)
    merged_cell[0][0].value = name_string + new_value
    
    
    
    # Writing Designation to the excel sheet Demo(Name: Assistant Professor)
    merged_cell = sheet[ 'A4:C4' ]
    new_value = bangla_teachers_designation[cteacher.designation]
    #print("New Value is : ", new_value)
    merged_cell[0][0].value = podobi_string + new_value
    
    
    # Writing Department to the excel sheet Demo (Dept: CSE )
    
    merged_cell = sheet[ 'A5:B5' ]
    new_value = cteacher.dept
    dept_name_bengali = translate_to_bengali(new_value)
    merged_cell[0][0].value = bivag_string + dept_name_bengali
    
    # Question Paper Setter Info to excel
    
   
    
    # Writing BSC ENG Name to the excel file
    
    for row in sheet.iter_rows(min_row=9, max_row=30, min_col=3, max_col=3):
        for cell in row:
            cell.value = "বি. এসসি. ইঞ্জি:"
       
       
    # Question Niamon
        
    # cell = sheet.cell(row=11, column=7)
    # cell.value = 2  
    
    # Question Examinar
    cell = sheet.cell(row=12, column=7)
    cell.value = cteacher.question_paper_setter
    
    #Clss Tests
    cell = sheet.cell(row=14, column=7)
    cell.value = cteacher.examiners_class_tests
    
    
    # Seminar 
    cell = sheet.cell(row=16, column=7)
    cell.value = cteacher.seminar
    
    #Total Teacher Attending Seminars
    cell = sheet.cell(row=16, column=8)
    ck =size(seminar_map)
    cell.value=ck
    
    #Sessionals
    cell = sheet.cell(row=17, column=7)
    cell.value = cteacher.examiners_sessional_classes
    
    y = cteacher.sessional_course
    cell = sheet.cell(row=17, column=8)
    if(y != "Null"):
        x = sessional_course_Cradit_map[y]
        
       
        #print("Teacher Name: ")
        #print(cteacher.name)
        
        cell.value =  float(x)
        #print(cell.value)
    
    
    # Subject Name
    
    cell = sheet.cell(row=12, column=5)
    z=cteacher.theory_course
    cell.value = z
    
    
    if(cell.value != "Null"):
        cell = sheet.cell(row=9, column=7)
    # Unmerge the cell before setting the value
        cell.value = 1
        
    
    #Class Test Subject
    cell = sheet.cell(row=14, column=5)
    z=cteacher.theory_course
    cell.value = z
    
    #Class Test Number
    
    cell = sheet.cell(row=14, column=8)
    cell.value = 1 
    
    
    
    #Sessional Course Subject
    cell = sheet.cell(row=17, column=5)
    z=cteacher.sessional_course
    cell.value=z
    
    
    #Central Viva (Total Teacher + Total Student)
    cell = sheet.cell(row=18, column=7)
    cell.value = cteacher.central_viva_voce
    
    cell = sheet.cell(row=18, column=8)
    ck=  size(examiners_sessional_classes_map)
    cell.value = ck
   
    
    
    
    #Thesis Progress Defense
    cell = sheet.cell(row=20, column=7)
    cell.value = cteacher.thesis_progress_defense
    
    
    cell = sheet.cell(row=20, column=8)
    ck = size(thesis_progress_defense_map)
    cell.value = ck
    
    
    
    cell = sheet.cell(row=24, column=7)
    cell.value = cteacher.tabulation
    
    
    #Scrutinizer
    cell = sheet.cell(row=25, column=7)
    cell.value = cteacher.script_scrutinizer
    
    
    # List of Duty
    cell = sheet.cell(row=26, column=7)
    cell.value = cteacher.list_of_duty
    
    
    #Typing & Drawing
    
    cell = sheet.cell(row=27, column=7)
    cell.value = cteacher.typing_and_drawing
    
    
    #Final Grade Sheet Verification
    
    cell = sheet.cell(row=28, column=7)
    cell.value= cteacher.final_grade_sheet_verification
    
    #Student Advising
    
    cell = sheet.cell(row=29, column=7)
    cell.value = cteacher.student_advising
    
    # Taka Amount
    
    cell = sheet.cell(row=32, column=9)
    print("Cell Value is: ")
    print(cell.value)
    
    
    
    merged_cell = sheet[ 'A32:E32' ]
    
    #merged_cell[0][0].value = taka
    
    
    
    
    file_name = f"{cteacher.name}.xlsx"
    workbook.save(f'D:\\Academic\\Output Files\\{file_name}') 
    
    #print(cteacher.designation)
    
    



def main():
    doc_path = 'D:\\Academic\\System Final\\Exam Bill Demo.docx'

    try:
        document = Document(doc_path)
        print_tables(document)
        
    except Exception as e:
        print(f"An error occurred: {e}")
        
        
        
    
    
    # print(ck)

    #ll=dept_map["Dr. M. M. A. Hashem"]
    #print("Check")
    #print(ll)
    # Printing each element on a new line using a loop
    #for element in TName:
      #print(element)




    #teachers_array = [
    #Teacher("John Doe", "Professor", "Mathematics", 2, 3, 1, 4, 5, 2, 1, 3, 2, 2, 1, 4, "Sessional_Math", "Theory_Math"),
    #Teacher("Jane Smith", "Associate Professor", "Physics", 1, 2, 3, 2, 4, 3, 2, 2, 1, 3, 3, 1, "Sessional_Physics", "Theory_Physics"),
    # Add more instances as needed
    #]

   # print(teachers_array[0].name);
   # for course, credit in sessional_course_Cradit_map.items():
    #   print(f"Course: {course}, Credit: {credit}")






    num_teachers = len(TName)
    
    ck = size(sessional_course_map)
    print("Sessional size is: ")
    print(ck)

# Creating instances and adding them to the array
    for i in range(num_teachers):
        teacher_instance = create_teacher(i)
        teachers_array.append(teacher_instance)

# Accessing and printing information from the array of teachers
    for teacher in teachers_array:

        #print(teacher.name)
        #print(teacher.sessional_course)
        excel_add(teacher)
        #print(teacher.to_dict())
    print("Code Executed Successfully")                
                  
                  
                  
                  
                  
                  
                  



if __name__ == "__main__":
    
    main()
    
    
    
    
    